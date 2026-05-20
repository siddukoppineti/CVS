#!/usr/bin/env python3

import csv
from openpyxl import load_workbook

MONITOR_FILE = "file_monitors.txt"     # change to file_monitor.txt if needed
EXCEL_FILE = "Varun_list.xlsx"
OUTPUT_FILE = "matched_monitors.csv"

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def normalize(value):
    return clean(value).lower()

def short_name(hostname):
    hostname = clean(hostname)
    return hostname.split(".", 1)[0].lower()

def find_col(headers, candidates):
    header_map = {normalize(h): i for i, h in enumerate(headers)}
    for c in candidates:
        c = normalize(c)
        if c in header_map:
            return header_map[c]
    return None

# Read monitor/server names
with open(MONITOR_FILE, "r", encoding="utf-8") as f:
    servers = [line.strip() for line in f if line.strip()]

# Load Excel workbook
wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
if not rows:
    raise SystemExit("Excel file is empty")

headers = list(rows[0])

name_idx = find_col(headers, ["Name"])
fqdn_idx = find_col(headers, ["Fully qualified domain name", "FQDN", "Fully Qualified Domain Name"])
app_idx = find_col(headers, ["Application Relationship"])
group_idx = find_col(headers, ["Support Group", "Assignment Group"])

if name_idx is None and fqdn_idx is None:
    raise SystemExit("Could not find Name or Fully qualified domain name column in Excel file")

# Build lookup
lookup = {}

for row in rows[1:]:
    if not row:
        continue

    name_val = clean(row[name_idx]) if name_idx is not None and name_idx < len(row) else ""
    fqdn_val = clean(row[fqdn_idx]) if fqdn_idx is not None and fqdn_idx < len(row) else ""
    app_val = clean(row[app_idx]) if app_idx is not None and app_idx < len(row) else ""
    group_val = clean(row[group_idx]) if group_idx is not None and group_idx < len(row) else ""

    if name_val:
        lookup[normalize(name_val)] = (app_val, group_val)
    if fqdn_val:
        lookup[normalize(fqdn_val)] = (app_val, group_val)
        lookup[short_name(fqdn_val)] = (app_val, group_val)

# Write output
with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["monitor", "server_name", "application_name", "assignment_group"])

    for server in servers:
        key = normalize(server)
        app = ""
        group = ""

        if key in lookup:
            app, group = lookup[key]
        else:
            sn = short_name(server)
            if sn in lookup:
                app, group = lookup[sn]

        writer.writerow([
            "file monitor",
            server,
            app if app else "MISSING",
            group if group else "MISSING"
        ])

print(f"Created {OUTPUT_FILE}")
